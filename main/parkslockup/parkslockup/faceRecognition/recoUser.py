import cv2
import face_recognition
from openpyxl import load_workbook
import numpy as np
import registUser



def execute():
    encodedFaces = loadEncodedImg()
    compareWithWebCam(encodedFaces)

def loadEncodedImg():
    wb = load_workbook('user.xlsx')
    ws = wb.active
    
    encodedFaces = []
    for row in ws.iter_rows(values_only=True):
        encodedFaces.append(np.array(row))
    # np.array 로 이루어진 배열
    return encodedFaces


def compareWithWebCam(encodedFaces):
    video_capture = cv2.VideoCapture(0)

    face_locations = []
    face_encodings = []
    face_names = []

    process_this_frame = True

    while True:
        ret, frame = video_capture.read()

        if process_this_frame:
            # 빠른 얼굴인식을 위해 4/1 배로 웹캠을 줄인다
            small_frame = cv2.resize(frame, (0,0), fx=0.25, fy=0.25)
            # BGR 로 들어온 색을 RGB로 바꾼다(BGR cv2에서 사용된다)
            rgb_small_frame = np.ascontiguousarray(small_frame[:, :, ::-1])

            # 웹캠에서 얼굴을 찾고 인코딩
            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

            face_names = []

            # 여러 개의 얼굴을 비교한다
            for face_encoding in face_encodings:

                # face_distance 는 얼굴 랜드마크간의 차이를 벡터로 나타낸다, 보통 한 0.45 정도 나오면 그 사람이다
                face_distances = face_recognition.face_distance(encodedFaces,face_encoding)
                print("faceDistance",face_distances)

                # np.argmin 으로 그 벡터값들 중 가장 작은값(차이가 적은값)을 가져온다
                best_match_index = np.argmin(face_distances)

                # 0.40 라면 그 사람일 확률이 높다
                name = "Unknown"
                if face_distances[best_match_index] < 0.4:
                    name = getFileNameFromRegisterlist(best_match_index)
                    
                face_names.append(name)

        # display 작업할 때는 faceCompare 작업이 실행되지 않도록 한다
        process_this_frame = not process_this_frame

        displayRectangle(frame,face_locations,face_names)

        # Display the resulting image
        cv2.imshow('Video', frame)

        # Hit 'q' on the keyboard to quit!
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release handle to the webcam
    video_capture.release()
    cv2.destroyAllWindows()

def getFileNameFromRegisterlist(index):
    withoutExten = registUser.registerList[index].split(".")[1]
    return withoutExten.split('/')[-1]


def displayRectangle(frame,face_locations,face_names):
     for (top, right, bottom, left), name in zip(face_locations, face_names):
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4

            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
            font = cv2.FONT_HERSHEY_DUPLEX
            namedUtf8 = name.encode('utf-8').decode('utf-8')
            print(name)
            print(namedUtf8)
            cv2.putText (frame, namedUtf8, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)